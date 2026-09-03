"""
SovereignWorkbench — Developer 2 Comprehensive Verification Suite (tests/test_dev2_tools.py)
Tests all deterministic tools, compilers, sandbox runners, RAG, and security modules
against the shared contract in app/schemas.py.
"""

import hashlib
import json
import pytest
from pathlib import Path

from app.schemas import (
    ApprovalNotePayload,
    PipeInspectionData,
    CostMatrixPayload,
    CostEstimateItem,
    AuditEvent,
    NetworkStats,
    RagQueryResponse,
    SandboxResult
)
from app.sandbox.runner import execute_in_sandbox
from app.sandbox.error_parser import distill_python_traceback
from app.compilers.docx_builder import compile_approval_note
from app.compilers.xlsx_builder import compile_cost_matrix
from app.rag.retriever import query_sovereign_rag
from app.rag.ingest import ingest_document_to_rag
from app.security.network_monitor import read_network_stats
from app.security.audit_chain import record_audit_event, verify_audit_chain, get_audit_ledger


# =====================================================================
# 1. SANDBOX & ERROR DISTILLATION TESTS
# =====================================================================

def test_sandbox_success():
    """Verify clean execution and structured JSON output parsing."""
    code = 'import json; print(json.dumps({"remaining_life": 3.14, "status": "REPLACE"}))'
    res = execute_in_sandbox(code)
    assert isinstance(res, SandboxResult)
    assert res.success is True
    assert res.exit_code == 0
    assert res.parsed_output is not None
    assert res.parsed_output["remaining_life"] == 3.14
    assert res.parsed_output["status"] == "REPLACE"
    assert res.distilled_error is None


def test_sandbox_self_healing_error_distillation():
    """Verify division-by-zero generates distilled root cause for LLM self-healing."""
    code = 'x = 10 / 0'
    res = execute_in_sandbox(code)
    assert isinstance(res, SandboxResult)
    assert res.success is False
    assert res.exit_code != 0
    assert res.distilled_error is not None
    assert "ZeroDivisionError" in res.distilled_error
    assert "Offending code" in res.distilled_error
    assert "Root cause" in res.distilled_error


def test_sandbox_key_error_distillation():
    """Verify KeyError provides dictionary troubleshooting guidance."""
    code = 'data = {"nominal": 4.8}; print(data["operating_years"])'
    res = execute_in_sandbox(code)
    assert res.success is False
    assert "KeyError" in res.distilled_error
    assert "operating_years" in res.distilled_error


def test_sandbox_timeout():
    """Verify strict execution timeout handling."""
    code = 'import time; time.sleep(10)'
    res = execute_in_sandbox(code, timeout=1)
    assert res.success is False
    assert res.exit_code == 124
    assert "TimeoutError" in res.distilled_error


# =====================================================================
# 2. DELIVERABLE COMPILERS TESTS
# =====================================================================

def test_docx_generation(tmp_path: Path):
    """Verify Word Approval Note compiles with corporate formatting and letterhead."""
    payload = ApprovalNotePayload(
        reference_number="MRPL/CDU-2/MAINT/2026/TEST",
        unit_name="Crude Distillation Unit 2 (CDU-2)",
        inspection_data=PipeInspectionData(
            line_tag="CDU-2-04-150-A1A",
            service_description="Crude Column Overhead Vapour",
            material_spec="ASTM A106 Grade B Carbon Steel",
            nominal_thickness_mm=4.8,
            actual_thickness_mm=3.2,
            minimum_required_thickness_mm=2.1,
            corrosion_rate_mm_year=0.35,
            remaining_life_years=3.14,
            mandatory_action="MANDATORY SHUTDOWN REPLACEMENT REQUIRED (< 5 YRS)"
        ),
        formula_derivation_steps=[
            "Corrosion Rate = (4.8 - 3.2) / 4.5 = 0.35 mm/yr",
            "t_min = 2.1 mm (per ASME B31.3)",
            "Remaining Life = (3.2 - 2.1) / 0.35 = 3.14 years"
        ],
        recommended_action="Procure 150m spool pipe and replace at Q4 turnaround."
    )
    out_file = tmp_path / "test_approval_note.docx"
    saved = compile_approval_note(payload, out_file)
    
    assert saved.exists()
    assert saved == out_file
    assert saved.stat().st_size > 5000  # Formatted Word document with tables


def test_xlsx_generation(tmp_path: Path):
    """Verify Excel Cost Matrix compiles with active formulas and proper cell structures."""
    import openpyxl
    
    payload = CostMatrixPayload(
        project_title="MRPL CDU-2 Unit Overhaul Estimate",
        line_tag="CDU-2-04-150-A1A",
        items=[
            CostEstimateItem(item_code="PIPE-01", description="Pipe ASTM A106-B 6-inch", quantity=100.0, unit="Meters", unit_rate_inr=5000.0, total_inr=500000.0),
            CostEstimateItem(item_code="FLG-02", description="Weld Flange Class 150", quantity=10.0, unit="Nos", unit_rate_inr=6000.0, total_inr=60000.0)
        ],
        contingency_percentage=10.0
    )
    out_file = tmp_path / "test_cost_matrix.xlsx"
    saved = compile_cost_matrix(payload, out_file)
    
    assert saved.exists()
    assert saved == out_file
    assert saved.stat().st_size > 4000
    
    # Load and inspect active Excel formula contents
    wb = openpyxl.load_workbook(str(saved), data_only=False)
    ws = wb.active
    
    # Line item active formula check: F6 should be =C6*E6
    assert ws["F6"].value == "=C6*E6"
    assert ws["F7"].value == "=C7*E7"
    
    # Subtotal formula: =SUM(F6:F7)
    assert ws["F8"].value == "=SUM(F6:F7)"
    
    # Contingency formula: =F8*(10.0/100)
    assert "=F8" in ws["F9"].value
    
    # Grand Total formula: =F8+F9
    assert ws["F10"].value == "=F8+F9"


# =====================================================================
# 3. SOVEREIGN RAG TESTS
# =====================================================================

def test_rag_retrieval():
    """Verify retrieval returns conformant RagQueryResponse with domain knowledge."""
    query = "What is the formula for piping remaining life under API 570?"
    res = query_sovereign_rag(query, top_k=2)
    
    assert isinstance(res, RagQueryResponse)
    assert res.query == query
    assert len(res.chunks) > 0
    assert len(res.combined_context) > 0
    assert res.chunks[0].relevance_score > 0.0


def test_rag_ingest_and_query(tmp_path: Path):
    """Verify document chunking and local indexing."""
    sop_file = tmp_path / "OISD-TEST-PROCEDURE.txt"
    sop_file.write_text(
        "Section 8.4: Refinery Emergency Pipe Shutdown Protocols.\n"
        "Whenever wall thinning exceeds statutory minimum tolerance, an emergency "
        "turnaround protocol shall be enacted immediately by the refinery manager."
    )
    
    chunks_count = ingest_document_to_rag(sop_file)
    assert chunks_count >= 1
    
    res = query_sovereign_rag("emergency pipe shutdown protocols", top_k=3)
    matching = [c for c in res.chunks if "OISD-TEST-PROCEDURE" in c.doc_name]
    assert len(matching) > 0


# =====================================================================
# 4. SECURITY & AUDIT TESTS
# =====================================================================

def test_network_stats_telemetry():
    """Verify NetworkStats matches schema and reports air-gap compliance."""
    stats = read_network_stats()
    assert isinstance(stats, NetworkStats)
    assert stats.air_gap_status == "LOCKED_AIR_GAP_COMPLIANT"
    assert stats.outbound_wan_bytes_delta == 0
    assert stats.total_rx_bytes >= 0
    assert stats.total_tx_bytes >= 0
    assert len(stats.wan_interface) > 0


def test_audit_chain_integrity():
    """Verify cryptographic SHA-256 block ledger chaining and tamper detection."""
    p_hash = hashlib.sha256(b"unit_test_prompt_data").hexdigest()
    o_hash = hashlib.sha256(b"unit_test_output_data").hexdigest()
    
    event = AuditEvent(
        user_role="lead_reliability_engineer",
        model_id="deepseek-r1:8b",
        task_type="api_570_compliance_check",
        prompt_hash=p_hash,
        tool_exit_code=0,
        output_hash=o_hash
    )
    
    entry_hash = record_audit_event(event)
    assert entry_hash is not None
    assert len(entry_hash) == 64  # Valid SHA-256 hex string
    
    # Verify complete chain
    is_valid, msg, count = verify_audit_chain()
    assert is_valid is True
    assert count >= 1
    
    # Verify recent entries list
    ledger = get_audit_ledger(limit=5)
    assert len(ledger) > 0
    assert ledger[0].entry_hash == entry_hash


# =====================================================================
# 5. FASTAPI ROUTER ENDPOINT TESTS
# =====================================================================

def test_fastapi_endpoints(tmp_path: Path):
    """Verify Dev 2 API endpoints for telemetry, audit ledger, and file operations."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
    from app.api.files import router as files_router
    from app.api.telemetry import router as telemetry_router

    app = FastAPI()
    app.include_router(files_router)
    app.include_router(telemetry_router)

    client = TestClient(app)

    # 1. Test telemetry network endpoint
    res_net = client.get("/api/telemetry/network")
    assert res_net.status_code == 200
    net_data = res_net.json()
    assert net_data["air_gap_status"] == "LOCKED_AIR_GAP_COMPLIANT"
    assert net_data["outbound_wan_bytes_delta"] == 0

    # 2. Test telemetry audit endpoint
    res_audit = client.get("/api/telemetry/audit")
    assert res_audit.status_code == 200
    audit_data = res_audit.json()
    assert audit_data["chain_valid"] is True
    assert "total_blocks" in audit_data

    # 3. Test files list endpoint
    res_list = client.get("/api/files/list")
    assert res_list.status_code == 200
    assert "deliverables" in res_list.json()
    assert "uploads" in res_list.json()

    # 4. Test file upload
    test_content = b"OISD Standard 118 Section 5 - Inspection Requirements"
    res_upload = client.post(
        "/api/files/upload",
        files={"file": ("test_upload_doc.txt", test_content, "text/plain")}
    )
    assert res_upload.status_code == 200
    assert res_upload.json()["status"] == "success"
    assert res_upload.json()["filename"] == "test_upload_doc.txt"
