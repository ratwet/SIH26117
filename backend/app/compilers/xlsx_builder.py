"""
SovereignWorkbench — Excel Cost Matrix Compiler (app/compilers/xlsx_builder.py)
Generates audit-ready financial estimates and cost workbooks with active Excel formulas,
contingency calculations, and statutory risk flags using openpyxl.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.schemas import CostMatrixPayload


# Palette Constants
COLOR_NAVY_HEADER = "1B365D"
COLOR_ZEBRA_LIGHT = "F4F7FA"
COLOR_WHITE = "FFFFFF"
COLOR_BORDER_GRAY = "D0D5DD"
COLOR_CRITICAL_BG = "FEE2E2"
COLOR_CRITICAL_FG = "991B1B"
COLOR_SAFE_BG = "DCFCE7"
COLOR_SAFE_FG = "166534"


def compile_cost_matrix(payload: CostMatrixPayload, output_path: Path) -> Path:
    """
    Compile a formatted financial cost estimate workbook with dynamic Excel formulas.
    
    Args:
        payload: CostMatrixPayload containing line items and contingency percentage.
        output_path: Destination path for the .xlsx file.
        
    Returns:
        Path: Resolved output path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost & Procurement Estimate"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Style definitions
    font_title = Font(name="Calibri", size=14, bold=True, color="123456")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="4B5563")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    
    fill_header = PatternFill(start_color=COLOR_NAVY_HEADER, end_color=COLOR_NAVY_HEADER, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA_LIGHT, end_color=COLOR_ZEBRA_LIGHT, fill_type="solid")
    
    border_thin = Border(
        left=Side(style="thin", color=COLOR_BORDER_GRAY),
        right=Side(style="thin", color=COLOR_BORDER_GRAY),
        top=Side(style="thin", color=COLOR_BORDER_GRAY),
        bottom=Side(style="thin", color=COLOR_BORDER_GRAY)
    )
    border_total = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # -------------------------------------------------------------
    # 1. Title Block
    # -------------------------------------------------------------
    ws["A1"] = "MANGALORE REFINERY AND PETROCHEMICALS LIMITED"
    ws["A1"].font = font_title
    
    ws["A2"] = f"Technical Evaluation & Procurement Cost Matrix — {payload.project_title}"
    ws["A2"].font = font_subtitle
    
    ws["A3"] = f"Line Tag: {payload.line_tag} | Currency: Indian Rupees (INR)"
    ws["A3"].font = font_subtitle
    
    # -------------------------------------------------------------
    # 2. Table Headers (Row 5)
    # -------------------------------------------------------------
    headers = [
        "Item Code",
        "Description",
        "Quantity",
        "Unit",
        "Unit Rate (INR)",
        "Total Cost (INR)"
    ]
    
    start_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3, 4] else (align_left if col_idx == 2 else align_right)
        cell.border = border_thin
        
    # -------------------------------------------------------------
    # 3. Line Items & Formulas (Rows 6 to 6 + N - 1)
    # -------------------------------------------------------------
    data_start = 6
    current_row = data_start
    
    for item in payload.items:
        # Col A: Item Code
        cA = ws.cell(row=current_row, column=1, value=item.item_code)
        cA.alignment = align_center
        
        # Col B: Description
        cB = ws.cell(row=current_row, column=2, value=item.description)
        cB.alignment = align_left
        
        # Col C: Quantity
        cC = ws.cell(row=current_row, column=3, value=item.quantity)
        cC.alignment = align_center
        cC.number_format = "#,##0.00" if isinstance(item.quantity, float) and not item.quantity.is_integer() else "#,##0"
        
        # Col D: Unit
        cD = ws.cell(row=current_row, column=4, value=item.unit)
        cD.alignment = align_center
        
        # Col E: Unit Rate
        cE = ws.cell(row=current_row, column=5, value=item.unit_rate_inr)
        cE.alignment = align_right
        cE.number_format = "₹ #,##0.00"
        
        # Col F: Total Cost using ACTIVE FORMULA: =C{row}*E{row}
        cF = ws.cell(row=current_row, column=6, value=f"=C{current_row}*E{current_row}")
        cF.alignment = align_right
        cF.font = font_bold
        cF.number_format = "₹ #,##0.00"
        
        # Apply zebra stripe and borders
        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border_thin
            if (current_row - data_start) % 2 == 1:
                cell.fill = fill_zebra
                
        current_row += 1
        
    data_end = current_row - 1
    
    # -------------------------------------------------------------
    # 4. Summary Totals (Subtotal, Contingency, Grand Total)
    # -------------------------------------------------------------
    # Subtotal Row
    subtotal_row = current_row
    ws.cell(row=subtotal_row, column=5, value="Subtotal (INR):").font = font_bold
    ws.cell(row=subtotal_row, column=5).alignment = align_right
    
    c_sub = ws.cell(row=subtotal_row, column=6, value=f"=SUM(F{data_start}:F{data_end})")
    c_sub.font = font_bold
    c_sub.alignment = align_right
    c_sub.number_format = "₹ #,##0.00"
    c_sub.border = border_thin
    
    # Contingency Row
    contingency_row = current_row + 1
    contingency_pct = payload.contingency_percentage
    ws.cell(row=contingency_row, column=5, value=f"Contingency ({contingency_pct:.1f}%):").font = font_regular
    ws.cell(row=contingency_row, column=5).alignment = align_right
    
    c_cont = ws.cell(row=contingency_row, column=6, value=f"=F{subtotal_row}*({contingency_pct}/100)")
    c_cont.font = font_regular
    c_cont.alignment = align_right
    c_cont.number_format = "₹ #,##0.00"
    c_cont.border = border_thin
    
    # Grand Total Row
    grand_total_row = current_row + 2
    ws.cell(row=grand_total_row, column=5, value="Grand Total (INR):").font = font_bold
    ws.cell(row=grand_total_row, column=5).alignment = align_right
    
    c_grand = ws.cell(row=grand_total_row, column=6, value=f"=F{subtotal_row}+F{contingency_row}")
    c_grand.font = font_bold
    c_grand.alignment = align_right
    c_grand.number_format = "₹ #,##0.00"
    c_grand.border = border_total
    
    # -------------------------------------------------------------
    # 5. Statutory Risk Flag Badge (Cell in Summary Box)
    # -------------------------------------------------------------
    badge_row = grand_total_row + 2
    ws.cell(row=badge_row, column=2, value="Statutory Risk Evaluation:").font = font_bold
    
    rem_life = getattr(payload, "remaining_life_years", None)
    if rem_life is not None and rem_life >= 5.0:
        badge_text = "IN-SERVICE MONITORING ACCEPTABLE"
        badge_fg = "155724"  # Dark green
        badge_bg = "D4EDDA"  # Light green
        note_text = f"Compliance Note: Remaining Life ({rem_life} Years) >= 5.0 Years indicates acceptable wall thickness under API 570."
    else:
        badge_text = "MANDATORY REPLACEMENT REQUIRED"
        badge_fg = COLOR_CRITICAL_FG
        badge_bg = COLOR_CRITICAL_BG
        life_str = f" ({rem_life} Years)" if rem_life is not None else ""
        note_text = f"Compliance Note: Remaining Life{life_str} < 5.0 Years requires scheduled procurement under API 570."

    badge_cell = ws.cell(row=badge_row, column=3, value=badge_text)
    badge_cell.font = Font(name="Calibri", size=10, bold=True, color=badge_fg)
    badge_cell.fill = PatternFill(start_color=badge_bg, end_color=badge_bg, fill_type="solid")
    badge_cell.alignment = align_center
    badge_cell.border = border_thin
    
    # Note on API 570 compliance
    ws.cell(row=badge_row + 1, column=2, value=note_text).font = font_subtitle
    
    # -------------------------------------------------------------
    # 6. Auto-fit column widths
    # -------------------------------------------------------------
    column_widths = {
        1: 16,  # Item Code
        2: 38,  # Description
        3: 14,  # Quantity
        4: 12,  # Unit
        5: 22,  # Unit Rate
        6: 24   # Total Cost
    }
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
    wb.save(str(output_path))
    return output_path
